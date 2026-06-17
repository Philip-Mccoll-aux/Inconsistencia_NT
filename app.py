import io
import unicodedata
import re
import difflib
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import streamlit as st

HEADER_SEARCH_ROWS = 10


# ── utilidades ────────────────────────────────────────────────────────────────

def normalize_id(val):
    if val is None:
        return ""
    s = str(val).strip()
    if s.endswith(".0") and s[:-2].lstrip("-").isdigit():
        s = s[:-2]
    return s


def is_empty(val):
    if val is None:
        return True
    s = str(val).strip()
    return s == "" or s.lower() == "nan"


def normalize_comuna(c):
    c = str(c).strip().upper()
    c = "".join(
        ch for ch in unicodedata.normalize("NFD", c)
        if unicodedata.category(ch) != "Mn"
    )
    return re.sub(r"\s+", " ", c)


def split_comunas(val):
    if is_empty(val):
        return set()
    tokens = re.split(r"[,;/]", str(val))
    return {normalize_comuna(t) for t in tokens if t.strip()}


def find_header_row(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= HEADER_SEARCH_ROWS:
            break
        names = [str(v).strip() if v is not None else "" for v in row]
        if "DOCUMENTO" in names and "ID PAÑO" in names:
            col_map = {name: j for j, name in enumerate(names) if name}
            return i, col_map
    return None, {}


def get_col(col_map, *candidates):
    for name in candidates:
        if name in col_map:
            return col_map[name]
    return None


def get_or_create_col(ws, col_map, header_row_num, col_name):
    if col_name in col_map:
        return col_map[col_name]
    new_idx = ws.max_column
    ws.cell(row=header_row_num, column=new_idx + 1, value=col_name)
    col_map[col_name] = new_idx
    return new_idx


def write_cell(ws, row_num, col_idx, value):
    ws.cell(row=row_num, column=col_idx + 1, value=value)


def read_cell(row, col_idx):
    if col_idx is None or col_idx >= len(row):
        return None
    return row[col_idx]


# ── documento objetivo ────────────────────────────────────────────────────────

def find_target_doc(ws, header_row, c_doc):
    pattern = re.compile(r"^EAF\s+(\d+)-2026$", re.IGNORECASE)
    best_num, best_doc = -1, None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row:
            continue
        val = str(row[c_doc] or "").strip()
        m = pattern.match(val)
        if m:
            num = int(m.group(1))
            if num > best_num:
                best_num, best_doc = num, val
    return best_doc


def is_revisable(row, c_tipo):
    return str(row[c_tipo] or "").strip().upper() in ("RE", "LD")


# ── carga BD Alimentadores ────────────────────────────────────────────────────

def load_bd_alimentadores(file):
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    if "Alimentadores" not in wb.sheetnames:
        return None, "La hoja 'Alimentadores' no existe en BD Alimentadores."
    ws = wb["Alimentadores"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None, "BD Alimentadores está vacía."
    header = [str(v).strip() if v is not None else "" for v in rows[0]]
    col = {n: i for i, n in enumerate(header) if n}

    required = {"ID UNICO", "ID Punto Control", "ID COORDINADO", "Comunas_25F"}
    missing = required - set(col)
    if missing:
        return None, f"Faltan columnas en BD Alimentadores: {missing}"

    by_punto_ctrl    = {}
    by_id_unico      = {}
    by_id_unico_full = {}
    by_barra_coord   = {}   # (barra, coord) -> set of id_unico
    all_comunas      = set()  # universo de comunas válidas

    for row in rows[1:]:
        id_unico    = normalize_id(row[col["ID UNICO"]])
        id_pc       = normalize_id(row[col["ID Punto Control"]])
        id_coord    = normalize_id(row[col["ID COORDINADO"]])
        comunas_raw = str(row[col["Comunas_25F"]] or "").strip()
        nombre      = str(row[col["Nombre Alimentador_25F"]] or "").strip() if "Nombre Alimentador_25F" in col else ""
        barra       = str(row[col["Barra_25F"]]             or "").strip() if "Barra_25F"             in col else ""
        sust        = str(row[col["Subestación"]]           or "").strip() if "Subestación"           in col else ""

        comunas_set = split_comunas(comunas_raw)
        all_comunas |= comunas_set

        by_punto_ctrl.setdefault(id_pc, set()).add(id_unico)
        by_id_unico[id_unico] = comunas_set
        by_id_unico_full[id_unico] = {
            "nombre":      nombre,
            "barra":       barra,
            "subestacion": sust,
            "comunas_set": comunas_set,
        }
        if id_pc and id_coord:
            by_barra_coord.setdefault((id_pc, id_coord), set()).add(id_unico)

    wb.close()
    return {
        "by_punto_ctrl":    by_punto_ctrl,
        "by_id_unico":      by_id_unico,
        "by_id_unico_full": by_id_unico_full,
        "by_barra_coord":   by_barra_coord,
        "all_comunas":      all_comunas,
    }, None


# ── fecha falla desde BD Interrupciones ──────────────────────────────────────

def get_fecha_falla(wb, target_doc):
    ws = None
    for name in ("BD Interrupciones", "BD interrupciones", "BD INTERRUPCIONES"):
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        return ""

    header_idx, c_fecha, c_doc_int = None, None, None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= HEADER_SEARCH_ROWS:
            break
        names = [str(v).strip() if v is not None else "" for v in row]
        if "FECHA Y HORA DE INICIO" in names:
            c_fecha    = names.index("FECHA Y HORA DE INICIO")
            c_doc_int  = names.index("DOCUMENTO") if "DOCUMENTO" in names else None
            header_idx = i
            break

    if c_fecha is None:
        c_fecha = 8

    fechas = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if header_idx is not None and i <= header_idx:
            continue
        if c_doc_int is not None and target_doc:
            if str(read_cell(row, c_doc_int) or "").strip() != target_doc:
                continue
        val = read_cell(row, c_fecha)
        if isinstance(val, datetime):
            fechas.append(val)

    return min(fechas).strftime("%d-%m-%Y") if fechas else ""


# ── PRE-REVISIÓN ──────────────────────────────────────────────────────────────

def _similar_comuna(unknown, all_comunas):
    matches = difflib.get_close_matches(unknown, all_comunas, n=1, cutoff=0.8)
    return matches[0] if matches else None


def _similar_nombre(nombre, candidatos, by_id_unico_full):
    """Intenta desempatar múltiples ID UNICO candidatos por similitud de nombre."""
    if not nombre:
        return None
    nombres_cands = {id_u: by_id_unico_full.get(id_u, {}).get("nombre", "") for id_u in candidatos}
    best_id, best_score = None, 0.0
    for id_u, nom in nombres_cands.items():
        if not nom:
            continue
        score = difflib.SequenceMatcher(None, nombre.upper(), nom.upper()).ratio()
        if score > best_score:
            best_score, best_id = score, id_u
    return best_id if best_score >= 0.6 else None


_ETAPA_MAP = {
    "PRE_DUPLICIDAD_AGREGACION":             "Agregación",
    "PRE_HORAS_NORMALIZACION_DISTINTAS":     "Agregación",
    "PRE_DATOS_INSUFICIENTES_TRIADA":        "ID único",
    "PRE_ID_UNICO_NO_RESUELTO":              "ID único",
    "PRE_ID_UNICO_MULTIPLES_CANDIDATOS":     "ID único",
    "PRE_ID_UNICO_INCONSISTENTE":            "ID único",
    "PRE_COMUNA_VACIA":                      "Comunas",
    "PRE_COMUNA_NO_VALIDA":                  "Comunas",
    "PRE_COMUNA_NO_VALIDA_POSIBLE_SIMIL":    "Comunas",
    "PRE_MULTIPLES_COMUNAS_EN_CELDA":        "Comunas",
    "PRE_HORA_DISPONIBILIDAD_BARRA_DISTINTA": "Disponibilidad barra",
}

_RESUMEN_MAP = {
    "PRE_DUPLICIDAD_AGREGACION":             "Duplicidad de agregación",
    "PRE_HORAS_NORMALIZACION_DISTINTAS":     "Horas de normalización distintas",
    "PRE_DATOS_INSUFICIENTES_TRIADA":        "Triada incompleta",
    "PRE_ID_UNICO_NO_RESUELTO":              "ID único no resuelto",
    "PRE_ID_UNICO_MULTIPLES_CANDIDATOS":     "Triada con múltiples candidatos",
    "PRE_ID_UNICO_INCONSISTENTE":            "ID único inconsistente con triada",
    "PRE_COMUNA_VACIA":                      "Fila sin comuna",
    "PRE_COMUNA_NO_VALIDA":                  "Comuna no válida",
    "PRE_COMUNA_NO_VALIDA_POSIBLE_SIMIL":    "Comuna no válida, posible similitud",
    "PRE_MULTIPLES_COMUNAS_EN_CELDA":        "Múltiples comunas en celda",
    "PRE_HORA_DISPONIBILIDAD_BARRA_DISTINTA": "Barra con más de una hora de disponibilidad",
}

_ACCION_MAP = {
    "PRE_DUPLICIDAD_AGREGACION":             "Revisar filas duplicadas del mismo alimentador/comuna.",
    "PRE_HORAS_NORMALIZACION_DISTINTAS":     "Revisar y homologar hora de normalización.",
    "PRE_DATOS_INSUFICIENTES_TRIADA":        "Completar triada (ID PAÑO / ID BARRA / ID COORDINADO).",
    "PRE_ID_UNICO_NO_RESUELTO":              "Verificar triada en BD Alimentadores.",
    "PRE_ID_UNICO_MULTIPLES_CANDIDATOS":     "Revisar nombre de alimentador y coincidencias en BD Alimentadores.",
    "PRE_ID_UNICO_INCONSISTENTE":            "Corregir CLAVE UNICA o verificar triada.",
    "PRE_COMUNA_VACIA":                      "Completar o revisar comuna informada.",
    "PRE_COMUNA_NO_VALIDA":                  "Revisar escritura de comuna.",
    "PRE_COMUNA_NO_VALIDA_POSIBLE_SIMIL":    "Revisar escritura de comuna.",
    "PRE_MULTIPLES_COMUNAS_EN_CELDA":        "Separar comunas en filas distintas si corresponde.",
    "PRE_HORA_DISPONIBILIDAD_BARRA_DISTINTA": "Revisar y homologar hora de disponibilidad de barra.",
}


def run_pre_revision(wb, ws, header_row, col_map, bd, hallazgos, target_doc):
    """
    Pre-revisiones B, C, D, D.1 y E.
    - Escribe 6 columnas REV_PRE_* resumidas por fila en BD Consolidado.
    - Crea hoja PRE_REVISION_RESUMEN con un registro por hallazgo único.
    """
    if not target_doc:
        return []

    c_doc     = col_map.get("DOCUMENTO")
    c_pano    = col_map.get("ID PAÑO")
    c_barra   = col_map.get("ID BARRA PUNTO DE CONTROL")
    c_coord   = col_map.get("ID COORDINADO AFECTADO")
    c_tipo    = col_map.get("TIPO CLIENTE")
    c_clave   = get_col(col_map, "ID UNICO", "CLAVE UNICA", "CLAVE ALIMENTADOR")
    c_comunas = col_map.get("Comunas")
    c_norm    = col_map.get("FECHA Y HORA NORMALIZACIÓN DE CONSUMO")
    c_disp    = col_map.get("FECHA Y HORA DISPONIBILIDAD DE LA BARRA")
    c_nombre  = col_map.get("NOMBRE ALIMENTADOR / CONSUMO")

    hr1 = header_row + 1
    cp_clasi  = get_or_create_col(ws, col_map, hr1, "REV_PRE_CLASIFICACION")
    cp_ids    = get_or_create_col(ws, col_map, hr1, "REV_PRE_ID_HALLAZGOS")
    cp_resumen = get_or_create_col(ws, col_map, hr1, "REV_PRE_RESUMEN")
    cp_filas  = get_or_create_col(ws, col_map, hr1, "REV_PRE_FILAS_RELACIONADAS")
    cp_triada = get_or_create_col(ws, col_map, hr1, "REV_PRE_TRIADA")
    cp_accion = get_or_create_col(ws, col_map, hr1, "REV_PRE_ACCION_SUGERIDA")

    all_comunas    = bd.get("all_comunas", set())
    by_barra_coord = bd.get("by_barra_coord", {})
    by_id_full     = bd.get("by_id_unico_full", {})

    # Recoger filas revisables
    data_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row:
            continue
        if str(read_cell(row, c_doc) or "").strip() != target_doc:
            continue
        if c_tipo is not None and not is_revisable(row, c_tipo):
            continue
        data_rows.append((i, row))

    # findings: list of dicts (one per unique finding)
    findings = []  # {clasif, unidad, filas, triada, detalle, accion, row_idxs}

    # ── B: duplicidad / horas distintas ───────────────────────────────────────
    agg_grupos = {}
    for i, row in data_rows:
        clave  = normalize_id(read_cell(row, c_clave)) if c_clave is not None else ""
        h_norm = read_cell(row, c_norm) if c_norm is not None else None
        tokens = split_comunas(read_cell(row, c_comunas)) if c_comunas is not None else set()
        if not tokens:
            tokens = {None}
        for tok in tokens:
            agg_grupos.setdefault((clave, tok), []).append((i, h_norm))

    for (clave, tok), entries in agg_grupos.items():
        if not clave or len(entries) <= 1:
            continue
        idxs  = [e[0] for e in entries]
        horas = {e[1] for e in entries if e[1] is not None}
        fila_nums = sorted([x + 1 for x in idxs])
        if len(horas) > 1:
            clasif = "PRE_HORAS_NORMALIZACION_DISTINTAS"
            horas_str = " / ".join(
                h.strftime("%H:%M") if isinstance(h, datetime) else str(h)
                for h in sorted(horas, key=str)
            )
            det = (f"La combinación {target_doc}+{clave}+{tok} presenta horas de normalización distintas. "
                   f"Filas involucradas: {fila_nums}. Horas observadas: {horas_str}.")
        else:
            clasif = "PRE_DUPLICIDAD_AGREGACION"
            det = (f"La combinación {target_doc}+{clave}+{tok} aparece en más de una fila. "
                   f"Filas involucradas: {fila_nums}.")
        unidad = f"{target_doc} + {clave} + {tok}"
        findings.append({
            "clasif": clasif, "unidad": unidad, "filas": fila_nums,
            "triada": "", "detalle": det,
            "accion": _ACCION_MAP[clasif], "row_idxs": idxs
        })

    # ── C: validación ID único vs triada ──────────────────────────────────────
    # Consolidar por triada única para no repetir hallazgo por cada fila
    triada_findings = {}  # triada_key -> finding dict  (para E también)
    for i, row in data_rows:
        pano  = normalize_id(read_cell(row, c_pano))  if c_pano  is not None else ""
        barra = normalize_id(read_cell(row, c_barra)) if c_barra is not None else ""
        coord = normalize_id(read_cell(row, c_coord)) if c_coord is not None else ""
        clave = normalize_id(read_cell(row, c_clave)) if c_clave is not None else ""
        nombre_fila = str(read_cell(row, c_nombre) or "").strip() if c_nombre is not None else ""
        triada_str = f"{pano}|{barra}|{coord}"

        if not barra or not coord:
            faltantes = []
            if not pano:  faltantes.append("ID PAÑO")
            if not barra: faltantes.append("ID BARRA")
            if not coord: faltantes.append("ID COORDINADO")
            key = ("PRE_DATOS_INSUFICIENTES_TRIADA", triada_str)
            if key not in triada_findings:
                det = f"No es posible resolver el ID UNICO porque falta: {' / '.join(faltantes)}."
                triada_findings[key] = {
                    "clasif": "PRE_DATOS_INSUFICIENTES_TRIADA", "unidad": triada_str,
                    "filas": [], "triada": triada_str, "detalle": det,
                    "accion": _ACCION_MAP["PRE_DATOS_INSUFICIENTES_TRIADA"], "row_idxs": []
                }
            triada_findings[key]["filas"].append(i + 1)
            triada_findings[key]["row_idxs"].append(i)
            continue

        candidatos = by_barra_coord.get((barra, coord), set())
        if not candidatos:
            key = ("PRE_ID_UNICO_NO_RESUELTO", triada_str)
            if key not in triada_findings:
                det = f"No fue posible resolver ID UNICO para la triada {triada_str}."
                triada_findings[key] = {
                    "clasif": "PRE_ID_UNICO_NO_RESUELTO", "unidad": triada_str,
                    "filas": [], "triada": triada_str, "detalle": det,
                    "accion": _ACCION_MAP["PRE_ID_UNICO_NO_RESUELTO"], "row_idxs": []
                }
            triada_findings[key]["filas"].append(i + 1)
            triada_findings[key]["row_idxs"].append(i)
        elif len(candidatos) == 1:
            esperado = next(iter(candidatos))
            if clave and clave != esperado:
                key = ("PRE_ID_UNICO_INCONSISTENTE", triada_str, clave)
                if key not in triada_findings:
                    det = (f"La triada {triada_str} apunta a ID UNICO esperado {esperado}, "
                           f"pero el borrador informa {clave}.")
                    triada_findings[key] = {
                        "clasif": "PRE_ID_UNICO_INCONSISTENTE",
                        "unidad": f"{triada_str} + {clave}",
                        "filas": [], "triada": triada_str, "detalle": det,
                        "accion": _ACCION_MAP["PRE_ID_UNICO_INCONSISTENTE"], "row_idxs": []
                    }
                triada_findings[key]["filas"].append(i + 1)
                triada_findings[key]["row_idxs"].append(i)
        else:
            # Múltiples candidatos — intentar desempatar por nombre
            desempate = _similar_nombre(nombre_fila, candidatos, by_id_full)
            if desempate and clave and clave != desempate:
                key = ("PRE_ID_UNICO_INCONSISTENTE", triada_str, clave)
                if key not in triada_findings:
                    det = (f"La triada {triada_str} apunta a ID UNICO esperado {desempate} "
                           f"(desempate por nombre), pero el borrador informa {clave}.")
                    triada_findings[key] = {
                        "clasif": "PRE_ID_UNICO_INCONSISTENTE",
                        "unidad": f"{triada_str} + {clave}",
                        "filas": [], "triada": triada_str, "detalle": det,
                        "accion": _ACCION_MAP["PRE_ID_UNICO_INCONSISTENTE"], "row_idxs": []
                    }
                triada_findings[key]["filas"].append(i + 1)
                triada_findings[key]["row_idxs"].append(i)
            elif not desempate:
                key = ("PRE_ID_UNICO_MULTIPLES_CANDIDATOS", triada_str)
                if key not in triada_findings:
                    det = (f"La triada {triada_str} resuelve más de un ID UNICO en BD Alimentadores. "
                           f"No fue posible desempatar por nombre de alimentador.")
                    triada_findings[key] = {
                        "clasif": "PRE_ID_UNICO_MULTIPLES_CANDIDATOS", "unidad": triada_str,
                        "filas": [], "triada": triada_str, "detalle": det,
                        "accion": _ACCION_MAP["PRE_ID_UNICO_MULTIPLES_CANDIDATOS"], "row_idxs": []
                    }
                triada_findings[key]["filas"].append(i + 1)
                triada_findings[key]["row_idxs"].append(i)

    findings.extend(triada_findings.values())

    # ── D / D.1: comunas válidas ───────────────────────────────────────────────
    # Agrupar por (clasif, token) para no repetir la misma comuna inválida
    comuna_findings = {}  # (clasif, tok) -> finding
    for i, row in data_rows:
        raw = read_cell(row, c_comunas) if c_comunas is not None else None
        triada_str = (f"{normalize_id(read_cell(row, c_pano)) if c_pano is not None else ''}"
                      f"|{normalize_id(read_cell(row, c_barra)) if c_barra is not None else ''}"
                      f"|{normalize_id(read_cell(row, c_coord)) if c_coord is not None else ''}")
        if is_empty(raw):
            key = ("PRE_COMUNA_VACIA", i)  # per row
            det = "La fila no informa comuna."
            findings.append({
                "clasif": "PRE_COMUNA_VACIA", "unidad": f"Fila {i+1}",
                "filas": [i + 1], "triada": triada_str, "detalle": det,
                "accion": _ACCION_MAP["PRE_COMUNA_VACIA"], "row_idxs": [i]
            })
        else:
            tokens = split_comunas(raw)
            if len(tokens) > 1:
                key = ("PRE_MULTIPLES_COMUNAS_EN_CELDA", i)
                det = ("La celda contiene más de una comuna. Se deben tratar como "
                       "tokens independientes para el análisis, sin modificar el valor original.")
                findings.append({
                    "clasif": "PRE_MULTIPLES_COMUNAS_EN_CELDA", "unidad": f"Fila {i+1}",
                    "filas": [i + 1], "triada": triada_str, "detalle": det,
                    "accion": _ACCION_MAP["PRE_MULTIPLES_COMUNAS_EN_CELDA"], "row_idxs": [i]
                })
            for tok in tokens:
                if tok not in all_comunas:
                    simil = _similar_comuna(tok, all_comunas) if all_comunas else None
                    if simil:
                        clasif = "PRE_COMUNA_NO_VALIDA_POSIBLE_SIMIL"
                        det = f"La comuna informada {tok} no existe. Posible coincidencia: {simil}."
                    else:
                        clasif = "PRE_COMUNA_NO_VALIDA"
                        det = f"La comuna informada {tok} no existe en el listado general de comunas válidas."
                    c_key = (clasif, tok)
                    if c_key not in comuna_findings:
                        comuna_findings[c_key] = {
                            "clasif": clasif, "unidad": tok,
                            "filas": [], "triada": "", "detalle": det,
                            "accion": _ACCION_MAP[clasif], "row_idxs": []
                        }
                    comuna_findings[c_key]["filas"].append(i + 1)
                    comuna_findings[c_key]["row_idxs"].append(i)

    findings.extend(comuna_findings.values())

    # ── E: hora única de disponibilidad de barra ───────────────────────────────
    disp_grupos = {}
    for i, row in data_rows:
        barra  = normalize_id(read_cell(row, c_barra)) if c_barra is not None else ""
        h_disp = read_cell(row, c_disp) if c_disp is not None else None
        disp_grupos.setdefault(barra, []).append((i, h_disp))

    for barra, entries in disp_grupos.items():
        horas = {e[1] for e in entries if e[1] is not None}
        if len(horas) > 1:
            fila_nums = sorted([e[0] + 1 for e in entries])
            det = (f"La barra {barra} presenta más de una hora de disponibilidad en el "
                   f"documento {target_doc}. Filas involucradas: {fila_nums}.")
            findings.append({
                "clasif": "PRE_HORA_DISPONIBILIDAD_BARRA_DISTINTA",
                "unidad": f"{target_doc} + Barra {barra}",
                "filas": fila_nums, "triada": "—", "detalle": det,
                "accion": _ACCION_MAP["PRE_HORA_DISPONIBILIDAD_BARRA_DISTINTA"],
                "row_idxs": [e[0] for e in entries]
            })

    # ── Asignar IDs correlativos ───────────────────────────────────────────────
    for n, f in enumerate(findings, 1):
        f["id"] = f"PRE-{n:03d}"

    # ── Mapa row_idx -> lista de findings que lo afectan ─────────────────────
    row_to_findings = {}
    for f in findings:
        for idx in f["row_idxs"]:
            row_to_findings.setdefault(idx, []).append(f)

    # ── Escribir columnas en BD Consolidado ────────────────────────────────────
    for i, row in data_rows:
        pano_v  = normalize_id(read_cell(row, c_pano))  if c_pano  is not None else ""
        barra_v = normalize_id(read_cell(row, c_barra)) if c_barra is not None else ""
        coord_v = normalize_id(read_cell(row, c_coord)) if c_coord is not None else ""
        triada_fila = f"{pano_v}|{barra_v}|{coord_v}"

        flist = row_to_findings.get(i, [])
        if flist:
            clasifs   = list(dict.fromkeys(f["clasif"]   for f in flist))
            ids       = list(dict.fromkeys(f["id"]       for f in flist))
            resumenes = list(dict.fromkeys(_RESUMEN_MAP.get(f["clasif"], f["clasif"]) for f in flist))
            filas_rel = list(dict.fromkeys(str(f["filas"]) for f in flist if f["filas"]))
            triadas   = list(dict.fromkeys(f["triada"] for f in flist if f["triada"] and f["triada"] != "—"))
            acciones  = list(dict.fromkeys(f["accion"] for f in flist))

            write_cell(ws, i + 1, cp_clasi,   "; ".join(clasifs))
            write_cell(ws, i + 1, cp_ids,     "; ".join(ids))
            write_cell(ws, i + 1, cp_resumen,  "; ".join(resumenes))
            write_cell(ws, i + 1, cp_filas,   "; ".join(filas_rel))
            write_cell(ws, i + 1, cp_triada,  "; ".join(triadas) if triadas else triada_fila)
            write_cell(ws, i + 1, cp_accion,  "; ".join(acciones))
        else:
            write_cell(ws, i + 1, cp_clasi,  "OK")
            write_cell(ws, i + 1, cp_triada, triada_fila)

    # ── Crear hoja PRE_REVISION_RESUMEN ───────────────────────────────────────
    sheet_name = "PRE_REVISION_RESUMEN"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws_pre = wb.create_sheet(sheet_name)

    hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    pre_headers = [
        "ID_HALLAZGO", "DOCUMENTO", "ETAPA", "CLASIFICACION",
        "UNIDAD_AFECTADA", "FILAS_RELACIONADAS", "TRIADA",
        "DETALLE", "ACCION_SUGERIDA", "ESTADO_REVISION"
    ]
    col_widths = [12, 18, 22, 38, 42, 30, 22, 80, 60, 16]
    for c, (h, w) in enumerate(zip(pre_headers, col_widths), 1):
        cell = ws_pre.cell(row=1, column=c, value=h)
        cell.fill = hfill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
        ws_pre.column_dimensions[cell.column_letter].width = w

    for rn, f in enumerate(findings, 2):
        ws_pre.cell(row=rn, column=1, value=f["id"])
        ws_pre.cell(row=rn, column=2, value=target_doc)
        ws_pre.cell(row=rn, column=3, value=_ETAPA_MAP.get(f["clasif"], "Otro"))
        ws_pre.cell(row=rn, column=4, value=f["clasif"])
        ws_pre.cell(row=rn, column=5, value=f["unidad"])
        ws_pre.cell(row=rn, column=6, value=", ".join(str(x) for x in f["filas"]))
        ws_pre.cell(row=rn, column=7, value=f["triada"])
        c8 = ws_pre.cell(row=rn, column=8, value=f["detalle"])
        c8.alignment = Alignment(wrap_text=True, vertical="top")
        ws_pre.cell(row=rn, column=9, value=f["accion"])
        ws_pre.cell(row=rn, column=10, value="")  # ESTADO_REVISION — vacío para ingeniería

    # ── Hallazgos para el reporte (uno por finding único) ────────────────────
    for f in findings:
        hallazgos.append({
            "HU": "PRE",
            "Tipo": f["clasif"],
            "Documento": target_doc,
            "Detalle": f"{f['id']} | {f['unidad']} | Filas: {f['filas']}"
        })

    return []


# ── HU-001 ────────────────────────────────────────────────────────────────────

def run_hu001(ws, header_row, col_map, hallazgos):
    required = [
        "DOCUMENTO", "ID PAÑO", "ID BARRA PUNTO DE CONTROL",
        "ID COORDINADO AFECTADO", "TIPO CLIENTE",
        "FECHA Y HORA DISPONIBILIDAD DE LA BARRA",
        "FECHA Y HORA NORMALIZACIÓN DE CONSUMO",
        "OBSERVACIONES", "Comunas", "Mecanismo de Recuperación",
    ]
    missing = [k for k in required if k not in col_map]
    if missing:
        return [f"[HU-001] Faltan columnas: {missing}"]

    c_doc   = col_map["DOCUMENTO"]
    c_pano  = col_map["ID PAÑO"]
    c_barra = col_map["ID BARRA PUNTO DE CONTROL"]
    c_coord = col_map["ID COORDINADO AFECTADO"]
    c_tipo  = col_map["TIPO CLIENTE"]
    c_u     = col_map["FECHA Y HORA DISPONIBILIDAD DE LA BARRA"]
    c_v     = col_map["FECHA Y HORA NORMALIZACIÓN DE CONSUMO"]
    c_w     = col_map["OBSERVACIONES"]
    c_ao    = col_map["Mecanismo de Recuperación"]
    c_aq    = get_col(col_map,
                      "Comentarios mecanismo de recuperación",
                      "Comentario mecanismo de recuperación")

    target_doc = find_target_doc(ws, header_row, c_doc)
    if not target_doc:
        hallazgos.append({"HU": "HU-001", "Tipo": "SIN_DOCUMENTO_OBJETIVO",
                          "Documento": "-", "Detalle": "No se encontró documento EAF XXX-2026."})
        return []

    c_clasi = get_or_create_col(ws, col_map, header_row + 1, "REV_RECUPERACION_CLASIFICACION")

    data_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row:
            continue
        if str(row[c_doc] or "").strip() != target_doc or not is_revisable(row, c_tipo):
            continue
        data_rows.append((i, row))

    for i, row in data_rows:
        u_val, v_val = row[c_u], row[c_v]
        w_val, ao_val = row[c_w], row[c_ao]
        aq_val = row[c_aq] if c_aq is not None else None

        if not isinstance(u_val, datetime) or not isinstance(v_val, datetime):
            clasi = "DATOS_INSUFICIENTES_FECHAS"
        elif v_val < u_val:
            clasi = ("NORMALIZACION_ANTES_DISPONIBILIDAD_SIN_RESPALDO"
                     if is_empty(w_val) and is_empty(ao_val) and is_empty(aq_val)
                     else "NORMALIZACION_ANTES_DISPONIBILIDAD_CON_RESPALDO")
        else:
            clasi = "OK"

        write_cell(ws, i + 1, c_clasi, clasi)
        if clasi != "OK":
            hallazgos.append({"HU": "HU-001", "Tipo": clasi, "Documento": target_doc,
                              "Detalle": f"Fila {i+1} | Paño={row[c_pano]} | Barra={row[c_barra]} | U={u_val} | V={v_val}"})
    return []


# ── HU-002 ────────────────────────────────────────────────────────────────────

def run_hu002(ws, header_row, col_map, bd, hallazgos, hu_results):
    required = ["DOCUMENTO", "ID PAÑO", "ID BARRA PUNTO DE CONTROL",
                "ID COORDINADO AFECTADO", "TIPO CLIENTE"]
    missing = [k for k in required if k not in col_map]
    if missing:
        return [f"[HU-002] Faltan columnas: {missing}"]

    c_doc        = col_map["DOCUMENTO"]
    c_barra      = col_map["ID BARRA PUNTO DE CONTROL"]
    c_tipo       = col_map["TIPO CLIENTE"]
    c_id_unico   = get_col(col_map, "ID UNICO", "CLAVE UNICA", "CLAVE ALIMENTADOR")
    c_comentario = get_col(col_map,
                           "Comentario Distribuidora Inconsistencia Alimentadores",
                           "COMENTARIO DISTRIBUIDORA SOBRE ALIMENTADORES")

    target_doc = find_target_doc(ws, header_row, c_doc)
    if not target_doc:
        hallazgos.append({"HU": "HU-002", "Tipo": "SIN_DOCUMENTO_OBJETIVO",
                          "Documento": "-", "Detalle": "No se encontró documento EAF XXX-2026."})
        return []

    c_clasi = get_or_create_col(ws, col_map, header_row + 1, "REV_ALIMENTADORES_CLASIFICACION")
    grupos = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row:
            continue
        if str(row[c_doc] or "").strip() != target_doc or not is_revisable(row, c_tipo):
            continue
        doc, barra = normalize_id(row[c_doc]), normalize_id(row[c_barra])
        if is_empty(row[c_doc]) or is_empty(row[c_barra]):
            hallazgos.append({"HU": "HU-002", "Tipo": "DATOS_INSUFICIENTES", "Documento": doc,
                              "Detalle": f"Fila {i+1}: falta DOCUMENTO o ID BARRA."})
            write_cell(ws, i + 1, c_clasi, "DATOS_INSUFICIENTES")
            continue
        grupos.setdefault((doc, barra), []).append((i, row))

    for (doc, barra), filas in grupos.items():
        observados, comentario_dx = set(), ""
        for _, row in filas:
            if c_id_unico is not None:
                val = normalize_id(row[c_id_unico])
                if not is_empty(val):
                    observados.add(val)
            if c_comentario is not None and not is_empty(row[c_comentario]):
                comentario_dx = str(row[c_comentario]).strip()

        esperados = bd["by_punto_ctrl"].get(barra)
        if esperados is None:
            tipo, faltan, sobran = "BARRA_NO_EN_BASE", set(), set()
        elif not observados:
            tipo, faltan, sobran = "DATOS_INSUFICIENTES", set(), set()
        else:
            faltan = esperados - observados
            sobran = observados - esperados
            tipo = ("FALTAN_Y_SOBRAN_ALIMENTADORES" if faltan and sobran else
                    "FALTAN_ALIMENTADORES"           if faltan            else
                    "SOBRAN_ALIMENTADORES"           if sobran            else "OK")

        hu_results["alim"][(doc, barra)] = {
            "tipo": tipo, "faltan": faltan, "sobran": sobran, "filas": filas
        }

        for i, _ in filas:
            write_cell(ws, i + 1, c_clasi, tipo)

        if tipo != "OK":
            det = f"Barra {barra}"
            if faltan: det += f" | Faltan: {sorted(faltan)}"
            if sobran: det += f" | Sobran: {sorted(sobran)}"
            if comentario_dx: det += f" | Comentario Dx: {comentario_dx}"
            hallazgos.append({"HU": "HU-002", "Tipo": tipo, "Documento": doc, "Detalle": det})
    return []


# ── HU-003 ────────────────────────────────────────────────────────────────────

def run_hu003(ws, header_row, col_map, bd, hallazgos, hu_results):
    required = ["DOCUMENTO", "ID COORDINADO AFECTADO", "TIPO CLIENTE", "Comunas"]
    missing = [k for k in required if k not in col_map]
    if missing:
        return [f"[HU-003] Faltan columnas: {missing}"]

    c_doc        = col_map["DOCUMENTO"]
    c_tipo       = col_map["TIPO CLIENTE"]
    c_comunas    = col_map["Comunas"]
    c_id_unico   = get_col(col_map, "ID UNICO", "CLAVE UNICA", "CLAVE ALIMENTADOR")
    c_comentario = get_col(col_map,
                           "Comentario Distribuidora sobre Inconsistencia Comunas",
                           "COMENTARIOS DISTRIBUIDORA SOBRE INCONSISTENCIA DE COMUNAS")

    target_doc = find_target_doc(ws, header_row, c_doc)
    if not target_doc:
        hallazgos.append({"HU": "HU-003", "Tipo": "SIN_DOCUMENTO_OBJETIVO",
                          "Documento": "-", "Detalle": "No se encontró documento EAF XXX-2026."})
        return []

    c_clasi = get_or_create_col(ws, col_map, header_row + 1, "REV_COMUNAS_CLASIFICACION")
    grupos = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row:
            continue
        if str(row[c_doc] or "").strip() != target_doc or not is_revisable(row, c_tipo):
            continue
        doc = normalize_id(row[c_doc])
        if c_id_unico is None or is_empty(row[c_id_unico]):
            if is_empty(row[c_doc]) or is_empty(row[c_comunas]):
                hallazgos.append({"HU": "HU-003", "Tipo": "DATOS_INSUFICIENTES", "Documento": doc,
                                  "Detalle": f"Fila {i+1}: falta DOCUMENTO, ID UNICO o Comunas."})
                write_cell(ws, i + 1, c_clasi, "DATOS_INSUFICIENTES")
            continue
        id_unico = normalize_id(row[c_id_unico])
        grupos.setdefault((doc, id_unico), []).append((i, row))

    for (doc, id_unico), filas in grupos.items():
        comunas_obs, comentario_dx = set(), ""
        for _, row in filas:
            comunas_obs |= split_comunas(row[c_comunas])
            if c_comentario is not None and not is_empty(row[c_comentario]):
                comentario_dx = str(row[c_comentario]).strip()

        esperadas = bd["by_id_unico"].get(id_unico)
        if esperadas is None:
            tipo, faltan, sobran = "ID_NO_EN_BASE", set(), set()
        elif not comunas_obs:
            tipo, faltan, sobran = "DATOS_INSUFICIENTES", set(), set()
        else:
            faltan = esperadas - comunas_obs
            sobran = comunas_obs - esperadas
            tipo = ("FALTAN_Y_SOBRAN_COMUNAS" if faltan and sobran else
                    "FALTAN_COMUNAS"           if faltan            else
                    "SOBRAN_COMUNAS"           if sobran            else "OK")

        hu_results["comunas"][(doc, id_unico)] = {
            "tipo": tipo, "faltan": faltan, "sobran": sobran, "filas": filas
        }

        for i, _ in filas:
            write_cell(ws, i + 1, c_clasi, tipo)

        if tipo != "OK":
            det = f"ID UNICO {id_unico}"
            if faltan: det += f" | Faltan: {sorted(faltan)}"
            if sobran: det += f" | Sobran: {sorted(sobran)}"
            if comentario_dx: det += f" | Comentario Dx: {comentario_dx}"
            hallazgos.append({"HU": "HU-003", "Tipo": tipo, "Documento": doc, "Detalle": det})
    return []


# ── HU-004: comentarios CEN ───────────────────────────────────────────────────

def _comentario_mecanismos(filas, col_map):
    """
    Lista todos los alimentadores con normalización antes de disponibilidad,
    luego agrega el texto de solicitud UNA SOLA VEZ al final.
    """
    c_rec    = col_map.get("REV_RECUPERACION_CLASIFICACION")
    c_nombre = col_map.get("NOMBRE ALIMENTADOR / CONSUMO")
    c_sust   = get_col(col_map, "SUBESTACIÓN", "SUBESTACION")

    combos = []
    seen   = set()
    for _, row in filas:
        if str(read_cell(row, c_rec) or "").strip() == "NORMALIZACION_ANTES_DISPONIBILIDAD_SIN_RESPALDO":
            alim = str(read_cell(row, c_nombre) or "").strip()
            sust = str(read_cell(row, c_sust)   or "").strip()
            key  = (alim, sust)
            if key not in seen:
                seen.add(key)
                combos.append(key)

    if not combos:
        return ""

    lineas = [
        f"El alimentador {alim} ({sust}) tiene hora de normalización equivalente previa "
        f"a la hora de disponibilidad de la barra."
        for alim, sust in combos
    ]
    frase_cierre = (
        "Se solicita detallar el mecanismo de recuperación de estas cargas para este "
        "alimentador donde esto se presente, y con el desglose requerido por la SEC, "
        "esto es comunal y por zona de tarificación."
    )
    return "\n\n".join(lineas) + "\n\n" + frase_cierre


def _comentario_alimentadores(filas, col_map, hu_results, bd):
    c_alim      = col_map.get("REV_ALIMENTADORES_CLASIFICACION")
    c_barra     = col_map.get("ID BARRA PUNTO DE CONTROL")
    c_doc_col   = col_map.get("DOCUMENTO")
    c_nombre    = col_map.get("NOMBRE ALIMENTADOR / CONSUMO")
    c_barra_nom = col_map.get("Barra")
    c_sust      = get_col(col_map, "SUBESTACIÓN", "SUBESTACION")
    c_id_unico  = get_col(col_map, "ID UNICO", "CLAVE UNICA", "CLAVE ALIMENTADOR")

    TIPOS = {"SOBRAN_ALIMENTADORES", "FALTAN_ALIMENTADORES", "FALTAN_Y_SOBRAN_ALIMENTADORES"}
    barras_con_hallazgo = set()
    for _, row in filas:
        if str(read_cell(row, c_alim) or "").strip() in TIPOS:
            barras_con_hallazgo.add(
                (normalize_id(read_cell(row, c_doc_col)), normalize_id(read_cell(row, c_barra)))
            )

    sobran_lines, faltan_lines = set(), set()
    tiene_sobran = tiene_faltan = False

    for (doc, barra) in barras_con_hallazgo:
        res    = hu_results["alim"].get((doc, barra), {})
        tipo   = res.get("tipo", "")
        sobran = res.get("sobran", set())
        faltan = res.get("faltan", set())

        if tipo in ("SOBRAN_ALIMENTADORES", "FALTAN_Y_SOBRAN_ALIMENTADORES") and sobran:
            tiene_sobran = True
            for _, fila in res.get("filas", []):
                id_u = normalize_id(read_cell(fila, c_id_unico)) if c_id_unico is not None else ""
                if id_u in sobran:
                    sobran_lines.add(
                        f"Alimentador {str(read_cell(fila, c_nombre) or '').strip()} — "
                        f"Barra {str(read_cell(fila, c_barra_nom) or '').strip()} — "
                        f"S/E {str(read_cell(fila, c_sust) or '').strip()}"
                    )

        if tipo in ("FALTAN_ALIMENTADORES", "FALTAN_Y_SOBRAN_ALIMENTADORES") and faltan:
            tiene_faltan = True
            for id_u in faltan:
                info = bd.get("by_id_unico_full", {}).get(id_u, {})
                faltan_lines.add(
                    f"Alimentador {info.get('nombre', id_u)} — "
                    f"Barra {info.get('barra', '')} — "
                    f"S/E {info.get('subestacion', '')}"
                )

    if not tiene_sobran and not tiene_faltan:
        return ""

    bloques = []
    if tiene_sobran and sobran_lines:
        bloques.append(
            "Se identifican alimentadores que, conforme a las bases de datos revisadas, "
            "se encontrarían incorporados en forma excedente:\n\n"
            + "\n".join(sorted(sobran_lines))
        )
    if tiene_faltan and faltan_lines:
        bloques.append(
            "Se identifican alimentadores que, conforme a las bases de datos revisadas, "
            "pertenecerían a las barras y/o S/E informadas, pero no se encontrarían "
            "incorporados en los antecedentes remitidos:\n\n"
            + "\n".join(sorted(faltan_lines))
        )

    if tiene_sobran and tiene_faltan:
        frase = "Se solicita revisar, corregir o justificar su inclusión u omisión, incorporando el desglose comunal y por zona de tarificación requerido por la SEC."
    elif tiene_sobran:
        frase = "Se solicita revisar, corregir o justificar su inclusión, incorporando el desglose comunal y por zona de tarificación requerido por la SEC."
    else:
        frase = "Se solicita revisar, corregir o justificar su omisión, incorporando el desglose comunal y por zona de tarificación requerido por la SEC."

    return "\n\n".join(bloques) + "\n\n" + frase


def _comentario_comunas(filas, col_map, hu_results, bd):
    c_com       = col_map.get("REV_COMUNAS_CLASIFICACION")
    c_doc_col   = col_map.get("DOCUMENTO")
    c_nombre    = col_map.get("NOMBRE ALIMENTADOR / CONSUMO")
    c_barra_nom = col_map.get("Barra")
    c_id_unico  = get_col(col_map, "ID UNICO", "CLAVE UNICA", "CLAVE ALIMENTADOR")

    TIPOS = {"SOBRAN_COMUNAS", "FALTAN_COMUNAS", "FALTAN_Y_SOBRAN_COMUNAS"}
    id_unicos_con_hallazgo = set()
    for _, row in filas:
        if str(read_cell(row, c_com) or "").strip() in TIPOS:
            id_u = normalize_id(read_cell(row, c_id_unico)) if c_id_unico is not None else ""
            doc  = normalize_id(read_cell(row, c_doc_col))
            if id_u:
                id_unicos_con_hallazgo.add((doc, id_u))

    sobran_lines, faltan_lines = set(), set()
    tiene_sobran = tiene_faltan = False

    for (doc, id_u) in id_unicos_con_hallazgo:
        res    = hu_results["comunas"].get((doc, id_u), {})
        tipo   = res.get("tipo", "")
        sobran = res.get("sobran", set())
        faltan = res.get("faltan", set())
        g_filas = res.get("filas", [])

        alim_b  = next((str(read_cell(r, c_nombre)    or "").strip() for _, r in g_filas if read_cell(r, c_nombre)),    "")
        barra_b = next((str(read_cell(r, c_barra_nom) or "").strip() for _, r in g_filas if read_cell(r, c_barra_nom)), "")

        if tipo in ("SOBRAN_COMUNAS", "FALTAN_Y_SOBRAN_COMUNAS") and sobran:
            tiene_sobran = True
            for com in sorted(sobran):
                sobran_lines.add(f"Comuna {com} — Alimentador {alim_b} — Barra {barra_b}")

        if tipo in ("FALTAN_COMUNAS", "FALTAN_Y_SOBRAN_COMUNAS") and faltan:
            tiene_faltan = True
            info = bd.get("by_id_unico_full", {}).get(id_u, {})
            for com in sorted(faltan):
                faltan_lines.add(
                    f"Comuna {com} — Alimentador {info.get('nombre', alim_b)} — Barra {info.get('barra', barra_b)}"
                )

    if not tiene_sobran and not tiene_faltan:
        return ""

    bloques = []
    if tiene_sobran and sobran_lines:
        bloques.append(
            "Se identifican comunas que, conforme a las bases de datos revisadas, "
            "se encontrarían incorporadas en forma excedente para los alimentadores "
            "y/o barras informadas:\n\n"
            + "\n".join(sorted(sobran_lines))
        )
    if tiene_faltan and faltan_lines:
        bloques.append(
            "Se identifican comunas que, conforme a las bases de datos revisadas, "
            "pertenecerían a los alimentadores y/o barras informadas, pero no se "
            "encontrarían incorporadas en los antecedentes remitidos:\n\n"
            + "\n".join(sorted(faltan_lines))
        )

    if tiene_sobran and tiene_faltan:
        frase = "Se solicita revisar, corregir o justificar su inclusión u omisión, incorporando el desglose comunal y por zona de tarificación requerido por la SEC."
    elif tiene_sobran:
        frase = "Se solicita revisar, corregir o justificar su inclusión, incorporando el desglose comunal y por zona de tarificación requerido por la SEC."
    else:
        frase = "Se solicita revisar, corregir o justificar su omisión, incorporando el desglose comunal y por zona de tarificación requerido por la SEC."

    return "\n\n".join(bloques) + "\n\n" + frase


def build_comentario_cen(filas, col_map, hu_results, bd):
    parts = [
        _comentario_mecanismos(filas, col_map),
        _comentario_alimentadores(filas, col_map, hu_results, bd),
        _comentario_comunas(filas, col_map, hu_results, bd),
    ]
    return "\n\n".join(p for p in parts if p)


def run_hu004(wb, ws, header_row, col_map, bd, hu_results, target_doc, hallazgos):
    req_cols = ["REV_RECUPERACION_CLASIFICACION",
                "REV_ALIMENTADORES_CLASIFICACION",
                "REV_COMUNAS_CLASIFICACION"]
    missing = [c for c in req_cols if c not in col_map]
    if missing:
        return [f"[HU-004] Ejecute primero HU-001/002/003. Faltan columnas: {missing}"]

    if not target_doc:
        hallazgos.append({"HU": "HU-004", "Tipo": "SIN_DOCUMENTO_OBJETIVO",
                          "Documento": "-", "Detalle": "No se encontró documento EAF XXX-2026."})
        return []

    c_doc   = col_map["DOCUMENTO"]
    c_coord = col_map.get("Coordinado")
    c_sust  = get_col(col_map, "SUBESTACIÓN", "SUBESTACION")
    c_rec   = col_map["REV_RECUPERACION_CLASIFICACION"]
    c_alim  = col_map["REV_ALIMENTADORES_CLASIFICACION"]
    c_com   = col_map["REV_COMUNAS_CLASIFICACION"]

    fecha_falla = get_fecha_falla(wb, target_doc)

    grupos = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row:
            continue
        if str(read_cell(row, c_doc) or "").strip() != target_doc:
            continue
        rec_v  = str(read_cell(row, c_rec)  or "").strip()
        alim_v = str(read_cell(row, c_alim) or "").strip()
        com_v  = str(read_cell(row, c_com)  or "").strip()
        if not any(v and v != "OK" for v in (rec_v, alim_v, com_v)):
            continue
        coordinado = str(read_cell(row, c_coord) or "").strip() if c_coord is not None else ""
        key = (str(read_cell(row, c_doc) or "").strip(), coordinado)
        grupos.setdefault(key, []).append((i, row))

    sheet_name = "ANEXO_OBSERVACIONES"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws_anexo = wb.create_sheet(sheet_name)

    hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    headers = ["EAF", "Fecha falla", "Distribuidora Involucrada",
               "S/E con consumo afectado", "¿Envía información?",
               "Presenta formato solicitado", "Comentario CEN"]
    for c, h in enumerate(headers, 1):
        cell = ws_anexo.cell(row=1, column=c, value=h)
        cell.fill = hfill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    ws_anexo.column_dimensions["A"].width = 16
    ws_anexo.column_dimensions["B"].width = 14
    ws_anexo.column_dimensions["C"].width = 38
    ws_anexo.column_dimensions["D"].width = 30
    ws_anexo.column_dimensions["E"].width = 18
    ws_anexo.column_dimensions["F"].width = 22
    ws_anexo.column_dimensions["G"].width = 85

    if not grupos:
        ws_anexo.cell(row=2, column=1, value="Sin hallazgos para el documento objetivo.")
        return []

    for rn, ((doc, coordinado), filas) in enumerate(grupos.items(), 2):
        ses = sorted({str(read_cell(r, c_sust) or "").strip()
                      for _, r in filas if c_sust and read_cell(r, c_sust)} - {""})
        comentario = build_comentario_cen(filas, col_map, hu_results, bd)

        ws_anexo.cell(row=rn, column=1, value=doc)
        ws_anexo.cell(row=rn, column=2, value=fecha_falla)
        ws_anexo.cell(row=rn, column=3, value=coordinado)
        c4 = ws_anexo.cell(row=rn, column=4, value="\n".join(ses))
        c4.alignment = Alignment(wrap_text=True, vertical="top")
        ws_anexo.cell(row=rn, column=5, value="Parcial")
        ws_anexo.cell(row=rn, column=6, value="")
        c7 = ws_anexo.cell(row=rn, column=7, value=comentario)
        c7.alignment = Alignment(wrap_text=True, vertical="top")

        hallazgos.append({"HU": "HU-004", "Tipo": "FILA_GENERADA", "Documento": doc,
                          "Detalle": f"Coordinado: {coordinado} | S/E: {', '.join(ses)}"})
    return []


# ── reporte ───────────────────────────────────────────────────────────────────

def build_report(wb, hallazgos, target_doc):
    sheet_name = "Reporte Hallazgos"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for c, h in enumerate(["HU", "Tipo", "Documento", "Detalle"], 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hfill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for r, h in enumerate(hallazgos, 2):
        ws.cell(row=r, column=1, value=h["HU"])
        ws.cell(row=r, column=2, value=h["Tipo"])
        ws.cell(row=r, column=3, value=h["Documento"])
        ws.cell(row=r, column=4, value=h["Detalle"])
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 90
    if target_doc:
        ws.cell(row=1, column=6, value="Documento objetivo:")
        ws.cell(row=1, column=7, value=target_doc)


# ── UI Streamlit ──────────────────────────────────────────────────────────────

st.set_page_config(page_title="Validador NT - Inconsistencias", layout="wide")
st.title("Validador de Inconsistencias NT")
st.markdown("Carga los dos archivos y ejecuta las validaciones.")

col1, col2 = st.columns(2)
with col1:
    borrador_file = st.file_uploader("📄 Archivo borrador EAF (Excel)", type=["xlsx"], key="borrador")
with col2:
    bd_file = st.file_uploader("📋 BD Alimentadores (Excel)", type=["xlsx"], key="bd")

hu_opts = st.multiselect(
    "Validaciones a ejecutar",
    ["PRE – Pre-revisiones (B/C/D/E)",
     "HU-001 – Mecanismos de recuperación",
     "HU-002 – Alimentadores por barra",
     "HU-003 – Comunas por alimentador",
     "HU-004 – Tabla observaciones CEN"],
    default=["PRE – Pre-revisiones (B/C/D/E)",
             "HU-001 – Mecanismos de recuperación",
             "HU-002 – Alimentadores por barra",
             "HU-003 – Comunas por alimentador",
             "HU-004 – Tabla observaciones CEN"],
)

btn_col1, btn_col2 = st.columns([1, 6])
with btn_col1:
    run_btn = st.button("▶ Ejecutar validaciones", type="primary",
                        disabled=(borrador_file is None or bd_file is None))
with btn_col2:
    reset_btn = st.button("🗑 Limpiar resultados", disabled="results" not in st.session_state)

if reset_btn:
    st.session_state.pop("results", None)
    st.rerun()

if run_btn:
    with st.spinner("Procesando…"):
        bd, err = load_bd_alimentadores(bd_file)
        if err:
            st.error(err)
            st.stop()

        wb = openpyxl.load_workbook(borrador_file, data_only=True)
        if "BD Consolidado" not in wb.sheetnames:
            st.error("El borrador no contiene la hoja 'BD Consolidado'.")
            st.stop()

        ws = wb["BD Consolidado"]
        header_row, col_map = find_header_row(ws)
        if header_row is None:
            st.error("No se encontró fila de encabezado con 'DOCUMENTO' e 'ID PAÑO'.")
            st.stop()

        hallazgos  = []
        all_errors = []
        hu_results = {"alim": {}, "comunas": {}}
        opts_str   = " ".join(hu_opts)

        target_doc = find_target_doc(ws, header_row, col_map["DOCUMENTO"])

        # Pre-revisiones siempre primero
        if "PRE" in opts_str:
            all_errors.extend(run_pre_revision(wb, ws, header_row, col_map, bd, hallazgos, target_doc))

        if "HU-001" in opts_str:
            all_errors.extend(run_hu001(ws, header_row, col_map, hallazgos))
        if "HU-002" in opts_str:
            all_errors.extend(run_hu002(ws, header_row, col_map, bd, hallazgos, hu_results))
        if "HU-003" in opts_str:
            all_errors.extend(run_hu003(ws, header_row, col_map, bd, hallazgos, hu_results))

        if "HU-004" in opts_str:
            all_errors.extend(
                run_hu004(wb, ws, header_row, col_map, bd, hu_results, target_doc, hallazgos)
            )

        build_report(wb, hallazgos, target_doc)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

    st.session_state["results"] = {
        "hallazgos":  hallazgos,
        "all_errors": all_errors,
        "target_doc": target_doc,
        "buf":        buf.getvalue(),
    }

if "results" in st.session_state:
    res        = st.session_state["results"]
    hallazgos  = res["hallazgos"]
    all_errors = res["all_errors"]
    target_doc = res["target_doc"]
    buf_bytes  = res["buf"]

    for e in all_errors:
        st.error(e)

    if target_doc:
        st.info(f"Documento objetivo: **{target_doc}**")

    h_pre = [h for h in hallazgos if h["HU"] == "PRE"]
    hu1   = [h for h in hallazgos if h["HU"] == "HU-001"]
    hu2   = [h for h in hallazgos if h["HU"] == "HU-002"]
    hu3   = [h for h in hallazgos if h["HU"] == "HU-003"]
    hu4   = [h for h in hallazgos if h["HU"] == "HU-004"]

    st.success(
        f"Validación completa — "
        f"PRE: {len(h_pre)} | HU-001: {len(hu1)} | HU-002: {len(hu2)} | "
        f"HU-003: {len(hu3)} | HU-004: {len(hu4)} fila(s) generada(s)"
    )

    tab0, tab1, tab2, tab3, tab4 = st.tabs(
        ["PRE Pre-revisiones", "HU-001 Mecanismos", "HU-002 Alimentadores",
         "HU-003 Comunas", "HU-004 Tabla CEN"]
    )

    def show_table(tab, items, label):
        with tab:
            if not items:
                st.info(f"Sin hallazgos para {label}.")
            else:
                import pandas as pd
                st.dataframe(pd.DataFrame(items)[["Tipo", "Documento", "Detalle"]],
                             use_container_width=True, hide_index=True)

    show_table(tab0, h_pre, "PRE")
    show_table(tab1, hu1,   "HU-001")
    show_table(tab2, hu2,   "HU-002")
    show_table(tab3, hu3,   "HU-003")
    show_table(tab4, hu4,   "HU-004")

    st.download_button(
        label="⬇ Descargar Excel (clasificaciones + ANEXO_OBSERVACIONES)",
        data=buf_bytes,
        file_name="EAF_validado.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
